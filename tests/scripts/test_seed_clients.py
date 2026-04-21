"""Unit tests for scripts/seed_clients.py.

Pure-function tests only — no DB, no sheet IO.
"""

from __future__ import annotations

import pytest

from scripts import seed_clients as sc


# ---------------------------------------------------------------------------
# normalize_email
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("  Foo@Bar.com ", "foo@bar.com"),
        ("no-at-sign", None),
        ("", None),
        (None, None),
    ],
)
def test_normalize_email(raw, expected):
    assert sc.normalize_email(raw) == expected


# ---------------------------------------------------------------------------
# derive_status
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Active", "active"),
        ("Churn", "churned"),
        ("Churn (Aus)", "churned"),
        ("Paused", "paused"),
        ("Paused (Leave)", "paused"),
        ("Ghost", "ghost"),
        ("N/A", "active"),
        ("", "active"),
        (None, "active"),
        ("  active  ", "active"),
        ("Unknown label", "active"),
    ],
)
def test_derive_status(raw, expected):
    assert sc.derive_status(raw) == expected


# ---------------------------------------------------------------------------
# derive_tags
# ---------------------------------------------------------------------------


def test_derive_tags_promoter_usa():
    tags = sc.derive_tags("active", nps_standing="Promoter", is_aus=False)
    assert "promoter" in tags
    assert "at_risk" not in tags
    assert "aus" not in tags


def test_derive_tags_detractor_sets_at_risk_and_detractor():
    tags = sc.derive_tags("active", nps_standing="Detractor / At Risk", is_aus=False)
    assert "at_risk" in tags
    assert "detractor" in tags


def test_derive_tags_churned_and_aus():
    tags = sc.derive_tags("churned", nps_standing=None, is_aus=True)
    assert "churned" in tags
    assert "aus" in tags


def test_derive_tags_nps_trailing_whitespace_still_matches():
    # NPS Standing values in the real sheet carry trailing spaces; the
    # transform trims before comparison.
    tags = sc.derive_tags("active", nps_standing="Detractor / At Risk ", is_aus=False)
    assert "detractor" in tags
    assert "at_risk" in tags


def test_derive_tags_never_produces_owing_money():
    """Tag dropped — sheet Standing column is unreliable per data-hygiene doc."""
    tags = sc.derive_tags("active", nps_standing=None, is_aus=False)
    assert "owing_money" not in tags


def test_derive_tags_does_not_set_at_risk_from_absent_nps():
    """at_risk used to also fire on Standing containing 'At risk'.
    That half was dropped; only NPS Detractor signals at_risk now."""
    tags = sc.derive_tags("active", nps_standing=None, is_aus=False)
    assert "at_risk" not in tags


# ---------------------------------------------------------------------------
# parse_owner
# ---------------------------------------------------------------------------


def test_parse_owner_clean_match():
    parsed = sc.parse_owner("Lou")
    assert parsed.team_email == "lou@theaipartner.io"
    assert parsed.is_clean_match is True


def test_parse_owner_trailing_space_is_clean():
    # "Nico " in the real sheet — trims to a clean match
    parsed = sc.parse_owner("Nico ")
    assert parsed.team_email == "nico@theaipartner.io"
    assert parsed.is_clean_match is True


def test_parse_owner_messy_takes_first_name_and_keeps_raw():
    parsed = sc.parse_owner("Lou (Scott Chasing)")
    assert parsed.team_email == "lou@theaipartner.io"
    assert parsed.is_clean_match is False
    assert parsed.raw == "Lou (Scott Chasing)"


def test_parse_owner_arrow_pattern_takes_first_named():
    parsed = sc.parse_owner("Lou > Nico?")
    assert parsed.team_email == "lou@theaipartner.io"
    assert parsed.is_clean_match is False


def test_parse_owner_unmapped_returns_none_with_raw():
    parsed = sc.parse_owner("Aleks")
    assert parsed.team_email is None
    assert parsed.raw == "Aleks"


@pytest.mark.parametrize("raw", ["N/A", "", None])
def test_parse_owner_blankish_returns_none_clean(raw):
    parsed = sc.parse_owner(raw)
    if raw == "N/A":
        # Not a mappable owner, raw preserved for reporting
        assert parsed.team_email is None
        assert parsed.raw == "N/A"
    else:
        assert parsed.team_email is None
        assert parsed.is_clean_match is True
        assert parsed.raw is None


# ---------------------------------------------------------------------------
# load_sheet_rows
# ---------------------------------------------------------------------------


def test_load_sheet_rows_reads_lowercase_tabs(tmp_path):
    import openpyxl as _openpyxl
    xlsx = tmp_path / "sheet.xlsx"
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "usa"
    ws.append(["Customer Name", "Client Emails", "Status", "Owner"])
    ws.append(["Real Client", "rc@example.com", "Active", "Lou"])
    ws.append(["Client No Email", None, "Active", "Lou"])
    aus = wb.create_sheet("aus")
    aus.append(["Customer Name", "Client Emails", "Status", "Owner"])
    aus.append(["Aussie", "au@example.com", "Active", "Lou"])
    wb.save(xlsx)

    rows, sheets_used = sc.load_sheet_rows(xlsx)

    assert "usa" in sheets_used and "aus" in sheets_used
    # Both USA rows survive row-filtering (blank-name only is skipped);
    # the missing-email one reaches build_client_payload and surfaces there.
    assert len(rows) == 3
    assert {r.country for r in rows} == {"USA", "AUS"}


def test_load_sheet_rows_still_reads_legacy_tab_names(tmp_path):
    import openpyxl as _openpyxl
    xlsx = tmp_path / "sheet.xlsx"
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "USA TOTALS"
    ws.append(["Customer Name", "Client Emails", "Status", "Owner "])
    ws.append(["Real Client", "rc@example.com", "Active", "Lou"])
    aus = wb.create_sheet("AUS TOTALS")
    aus.append(["Customer Name", "Client Emails", "Status", "Owner "])
    wb.save(xlsx)

    rows, sheets_used = sc.load_sheet_rows(xlsx)

    assert "USA TOTALS" in sheets_used
    assert len(rows) == 1


# ---------------------------------------------------------------------------
# build_client_payload — skip-missing-email rule
# ---------------------------------------------------------------------------


def _row(**overrides):
    base = {
        "customer name": "Jane Doe",
        "client emails": "jane@example.com",
        "client phone no.": "+15551234567",
        "slack user id": None,
        "date": None,
        "status": "Active",
        "standing": None,
        "nps standing": None,
        "owner": "Lou",
        "contracted rev": 9000,
    }
    base.update(overrides)
    return base


def test_build_client_payload_happy_path():
    payload = sc.build_client_payload(_row(), country="USA", seeded_at_iso="2026-04-21")
    assert payload is not None
    assert payload["email"] == "jane@example.com"
    assert payload["full_name"] == "Jane Doe"
    assert payload["status"] == "active"
    assert payload["program_type"] == sc.PROGRAM_TYPE == "9k_consumer"
    assert payload["timezone"] == sc.DEFAULT_TIMEZONE == "America/New_York"
    assert payload["metadata"]["country"] == "USA"
    assert payload["metadata"]["seeded_at"] == "2026-04-21"
    assert set(payload["metadata"].keys()) == {
        "seed_source", "seeded_at", "country", "nps_standing", "owner_raw"
    }
    assert "standing" not in payload["metadata"]


def test_build_client_payload_omits_revenue_fields():
    """Revenue fields in the sheet are stale (per Scott) and must not leak
    into clients.metadata. The schema of the metadata blob is fixed to the
    six keys above; no contracted_revenue* keys ever."""
    row = _row(**{"contracted rev": 9000, "contracted rev aud": 12000})
    payload = sc.build_client_payload(row, country="USA", seeded_at_iso="2026-04-21")
    assert payload is not None
    assert "contracted_revenue_usd" not in payload["metadata"]
    assert "contracted_revenue" not in payload["metadata"]
    assert "contracted_revenue_currency" not in payload["metadata"]


def test_build_client_payload_skips_missing_email():
    assert sc.build_client_payload(
        _row(**{"client emails": None}), country="USA", seeded_at_iso="2026-04-21"
    ) is None
    assert sc.build_client_payload(
        _row(**{"client emails": "   "}), country="USA", seeded_at_iso="2026-04-21"
    ) is None


def test_build_client_payload_aus_tag_without_revenue():
    row = _row()
    payload = sc.build_client_payload(row, country="AUS", seeded_at_iso="2026-04-21")
    assert payload is not None
    assert payload["metadata"]["country"] == "AUS"
    assert "aus" in payload["tags"]


# ---------------------------------------------------------------------------
# apply_clients — update / insert / reactivate
# ---------------------------------------------------------------------------


def _fake_proposed(email="jane@example.com"):
    return {
        "email": email,
        "full_name": "Jane",
        "phone": None,
        "slack_user_id": None,
        "start_date": "2026-04-01",
        "status": "active",
        "program_type": "9k_consumer",
        "timezone": "America/New_York",
        "tags": ["promoter"],
        "metadata": {"seed_source": "test", "country": "USA"},
    }


def test_apply_clients_reactivates_archived_row(mocker):
    """When an existing row is archived and the email appears again in
    the proposed set, apply_clients updates in place AND clears
    archived_at. The reactivations counter reflects it."""
    archived_row = {
        "id": "c1",
        "email": "jane@example.com",
        "metadata": {"prior_key": "prior_val"},
        "archived_at": "2026-04-20T00:00:00+00:00",
    }
    mocker.patch(
        "scripts.seed_clients.fetch_existing_client_emails",
        return_value={"jane@example.com": archived_row},
    )
    fake_db = mocker.MagicMock()

    _, inserts, updates, reactivations = sc.apply_clients(
        fake_db, [_fake_proposed()]
    )

    assert inserts == 0
    assert updates == 1
    assert reactivations == 1
    update_payload = fake_db.table.return_value.update.call_args[0][0]
    assert update_payload["archived_at"] is None
    # metadata merge: prior key preserved, new key overrides on collision
    assert update_payload["metadata"]["prior_key"] == "prior_val"
    assert update_payload["metadata"]["country"] == "USA"


def test_apply_clients_inserts_when_no_existing_row(mocker):
    mocker.patch("scripts.seed_clients.fetch_existing_client_emails", return_value={})
    fake_db = mocker.MagicMock()
    fake_db.table.return_value.insert.return_value.execute.return_value.data = [
        {"id": "new-1"}
    ]

    email_to_id, inserts, updates, reactivations = sc.apply_clients(
        fake_db, [_fake_proposed()]
    )

    assert inserts == 1
    assert updates == 0
    assert reactivations == 0
    assert email_to_id["jane@example.com"] == "new-1"


# ---------------------------------------------------------------------------
# compute_archival_plan / apply_archival
# ---------------------------------------------------------------------------


def _mock_db_with_chain(mocker, existing_clients, channels=None, assignments=None):
    """Build a supabase-py-shaped mock that returns canned rows for the three
    SELECT paths compute_archival_plan touches."""
    existing_clients = existing_clients or []
    channels = channels or []
    assignments = assignments or []

    fake_db = mocker.MagicMock()

    def table(name):
        chain = mocker.MagicMock()
        if name == "clients":
            chain.select.return_value.is_.return_value.execute.return_value.data = existing_clients
        if name == "slack_channels":
            chain.select.return_value.in_.return_value.eq.return_value.execute.return_value.data = channels
            chain.update.return_value.in_.return_value.eq.return_value.execute.return_value.data = channels
        if name == "client_team_assignments":
            chain.select.return_value.in_.return_value.is_.return_value.execute.return_value.data = assignments
            chain.update.return_value.in_.return_value.is_.return_value.execute.return_value.data = assignments
        # clients update path
        chain.update.return_value.in_.return_value.execute.return_value.data = existing_clients
        return chain

    fake_db.table.side_effect = table
    return fake_db


def test_compute_archival_plan_identifies_non_proposed_clients(mocker):
    existing = [
        {"id": "c1", "email": "keep@example.com",  "full_name": "Keep Me",  "status": "active"},
        {"id": "c2", "email": "drop@example.com",  "full_name": "Drop Me",  "status": "churned"},
        {"id": "c3", "email": "drop2@example.com", "full_name": "Drop Too", "status": "churned"},
    ]
    channels = [{"id": "ch1"}, {"id": "ch2"}]           # two channels linked to archived clients
    assignments = [{"id": "a1"}]                         # one active assignment linked

    fake_db = _mock_db_with_chain(mocker, existing, channels, assignments)

    plan = sc.compute_archival_plan(fake_db, proposed_emails={"keep@example.com"})

    assert {row["email"] for row in plan.clients_to_archive} == {
        "drop@example.com", "drop2@example.com",
    }
    assert plan.expected_channel_archivals == 2
    assert plan.expected_assignment_unassignments == 1


def test_compute_archival_plan_empty_when_all_proposed(mocker):
    existing = [{"id": "c1", "email": "a@example.com", "full_name": "A", "status": "active"}]
    fake_db = _mock_db_with_chain(mocker, existing)

    plan = sc.compute_archival_plan(fake_db, proposed_emails={"a@example.com"})

    assert plan.clients_to_archive == []
    assert plan.expected_channel_archivals == 0
    assert plan.expected_assignment_unassignments == 0


def test_apply_archival_noop_on_empty_plan(mocker):
    fake_db = mocker.MagicMock()
    counts = sc.apply_archival(fake_db, sc.ArchivalPlan())
    assert counts == (0, 0, 0)
    fake_db.table.assert_not_called()


def test_apply_archival_executes_all_three_writes(mocker):
    plan = sc.ArchivalPlan(
        clients_to_archive=[
            {"id": "c1", "email": "drop@example.com", "full_name": "D", "status": "churned"},
            {"id": "c2", "email": "drop2@example.com", "full_name": "D2", "status": "churned"},
        ],
        expected_channel_archivals=1,
        expected_assignment_unassignments=1,
    )

    fake_db = _mock_db_with_chain(
        mocker,
        existing_clients=plan.clients_to_archive,
        channels=[{"id": "ch1"}],
        assignments=[{"id": "a1"}],
    )

    counts = sc.apply_archival(fake_db, plan)

    # clients_count uses the update response; channels and assignments
    # each reflect their returned row lists.
    assert counts == (2, 1, 1)


# ---------------------------------------------------------------------------
# apply_log_breakdowns
# ---------------------------------------------------------------------------


def test_apply_log_breakdowns_counts_status_journey_tags_and_owners(mocker):
    client_rows = [
        {"status": "active",  "journey_stage": "onboarding", "tags": ["promoter"]},
        {"status": "active",  "journey_stage": None,         "tags": ["promoter", "aus"]},
        {"status": "paused",  "journey_stage": None,         "tags": []},
    ]
    assignment_rows = [
        {"team_member_id": "tm-lou"},
        {"team_member_id": "tm-lou"},
        {"team_member_id": "tm-scott"},
    ]
    team_rows = [
        {"id": "tm-lou",   "full_name": "Lou Perez"},
        {"id": "tm-scott", "full_name": "Scott Wilson"},
    ]

    fake_db = mocker.MagicMock()

    def table(name):
        chain = mocker.MagicMock()
        if name == "clients":
            chain.select.return_value.is_.return_value.execute.return_value.data = client_rows
        if name == "client_team_assignments":
            chain.select.return_value.is_.return_value.execute.return_value.data = assignment_rows
        if name == "team_members":
            chain.select.return_value.in_.return_value.execute.return_value.data = team_rows
        return chain

    fake_db.table.side_effect = table

    out = sc.apply_log_breakdowns(fake_db)

    assert "Breakdowns across 3 active clients" in out
    assert "promoter" in out and "aus" in out
    assert "Lou Perez" in out and "Scott Wilson" in out
    assert "primary_csm assignments" in out


# ---------------------------------------------------------------------------
# Dry-run vs apply separation — the main() script short-circuits on no --apply
# ---------------------------------------------------------------------------


def _make_trivial_sheet(tmp_path):
    xlsx = tmp_path / "sheet.xlsx"
    import openpyxl as _openpyxl
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "USA TOTALS"
    ws.append(["Customer Name", "Client Emails", "Status", "Owner "])
    ws.append(["Test Client", "tc@example.com", "Active", "Lou"])
    aus = wb.create_sheet("AUS TOTALS")
    aus.append(["Customer Name", "Client Emails", "Status", "Owner "])
    wb.save(xlsx)
    return xlsx


def test_main_dry_run_does_not_call_apply_paths(mocker, tmp_path):
    xlsx = _make_trivial_sheet(tmp_path)

    fake_db = mocker.MagicMock()
    mocker.patch("scripts.seed_clients.get_client", return_value=fake_db)
    mocker.patch(
        "scripts.seed_clients.fetch_existing_client_emails", return_value={}
    )
    mocker.patch(
        "scripts.seed_clients.compute_archival_plan", return_value=sc.ArchivalPlan()
    )
    apply_clients_spy = mocker.patch("scripts.seed_clients.apply_clients")
    apply_archival_spy = mocker.patch("scripts.seed_clients.apply_archival")

    rc = sc.main(["--input", str(xlsx)])

    assert rc == 0
    apply_clients_spy.assert_not_called()
    apply_archival_spy.assert_not_called()


def test_main_apply_calls_apply_paths(mocker, tmp_path):
    xlsx = _make_trivial_sheet(tmp_path)

    fake_db = mocker.MagicMock()
    mocker.patch("scripts.seed_clients.get_client", return_value=fake_db)
    mocker.patch(
        "scripts.seed_clients.fetch_existing_client_emails", return_value={}
    )
    mocker.patch(
        "scripts.seed_clients.compute_archival_plan",
        return_value=sc.ArchivalPlan(
            clients_to_archive=[{"id": "old-1", "email": "x@x.com", "full_name": "X", "status": "churned"}],
            expected_channel_archivals=0,
            expected_assignment_unassignments=0,
        ),
    )
    apply_clients_spy = mocker.patch(
        "scripts.seed_clients.apply_clients",
        return_value=({"tc@example.com": "id-1"}, 1, 0, 0),
    )
    mocker.patch(
        "scripts.seed_clients.resolve_team_member_ids",
        return_value={"lou@theaipartner.io": "tm-lou"},
    )
    mocker.patch("scripts.seed_clients.apply_channels", return_value=0)
    mocker.patch("scripts.seed_clients.apply_assignments", return_value=1)
    apply_archival_spy = mocker.patch(
        "scripts.seed_clients.apply_archival", return_value=(1, 0, 0)
    )
    mocker.patch("scripts.seed_clients.apply_log_breakdowns", return_value="(breakdown)")
    mocker.patch("scripts.seed_clients.write_log")

    rc = sc.main(["--input", str(xlsx), "--apply"])

    assert rc == 0
    apply_clients_spy.assert_called_once()
    apply_archival_spy.assert_called_once()
