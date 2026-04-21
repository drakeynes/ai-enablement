"""Seed the `clients`, `slack_channels`, and `client_team_assignments`
tables from the Financial Master Sheet.

Usage:
    # Dry run (default) — prints the proposed changes, writes nothing.
    python scripts/seed_clients.py

    # Apply — writes the proposed changes, produces an import log.
    python scripts/seed_clients.py --apply

    # Alternate input path.
    python scripts/seed_clients.py --input path/to/sheet.xlsx

The script reads the USA TOTALS and AUS TOTALS tabs only. All other
tabs are ignored. See docs/runbooks/seed_clients.md for the full
workflow, common fixes, and the re-export loop.
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl

# Make sibling `shared` package importable when run as a script.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from shared.db import get_client  # noqa: E402

USA_TAB = "USA TOTALS"
AUS_TAB = "AUS TOTALS"
DEFAULT_INPUT_DIR = _REPO_ROOT / "data" / "client_seed"
IMPORT_LOG_DIR = _REPO_ROOT / "data" / "client_seed"

OWNER_FIRST_NAMES: dict[str, str] = {
    "lou": "lou@theaipartner.io",
    "scott": "scott@theaipartner.io",
    "nico": "nico@theaipartner.io",
    "nabeel": "nabeel@theaipartner.io",
    "aman": "aman@theaipartner.io",
}

STATUS_MAP: dict[str, str] = {
    "active": "active",
    "churn": "churned",
    "churn (aus)": "churned",
    "paused": "paused",
    "paused (leave)": "paused",
    "ghost": "ghost",
    "n/a": "active",
    "": "active",
}

# The Active++ working-view filter. Rows with a mapped status outside these
# sets are not imported. USA uses Scott's "Active++" saved view; AUS mirrors
# it minus ghost (AUS data has no ghosts today; the rule is future-proofing).
# See docs/runbooks/seed_clients.md for the canonical definition.
ACTIVE_PLUS_PLUS_BY_COUNTRY: dict[str, frozenset[str]] = {
    "USA": frozenset({"active", "ghost", "paused"}),
    "AUS": frozenset({"active", "paused"}),
}


def is_in_active_plus_plus_view(status: str, country: str) -> bool:
    """True when the mapped status is allowed for the row's source tab."""
    return status in ACTIVE_PLUS_PLUS_BY_COUNTRY.get(country, frozenset())

# Customer Name values that are sheet-aggregate rows, not real clients.
# Filtered before the missing-email check so they don't pollute the skipped
# report. Match is case-insensitive on the trimmed string.
AGGREGATE_ROW_LABELS: frozenset[str] = frozenset(
    label.lower()
    for label in (
        "TOTALS",
        "UF Collection Rate",
        "BE Collection Opportunity",
        "BE Collections Rate",
        "BE Owining Rate",
        "Upsell Rate",
        "Referral Rate",
        "Refund Rate",
        "Total Active Clients",
        "Referrals",
        "Upsells",
        "Retention",
    )
)


def is_aggregate_row(customer_name: Any) -> bool:
    """True when Customer Name is a known sheet-aggregate label."""
    if customer_name is None:
        return False
    return str(customer_name).strip().lower() in AGGREGATE_ROW_LABELS


# ---------------------------------------------------------------------------
# Pure transforms — no IO, trivially testable.
# ---------------------------------------------------------------------------


def normalize_email(raw: Any) -> str | None:
    """Return a lowercased, trimmed email or None when blank / non-email."""
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text or "@" not in text:
        return None
    return text


def derive_status(raw: Any) -> str:
    """Map a sheet Status value to the `clients.status` vocabulary."""
    if raw is None:
        return "active"
    key = str(raw).strip().lower()
    return STATUS_MAP.get(key, "active")


def derive_tags(
    status: str,
    nps_standing: str | None,
    is_aus: bool,
) -> list[str]:
    """Derive the `clients.tags` array from sheet-side signals.

    Only NPS-derived signals and stable attributes (country, status)
    are used. The Standing column was previously a source but its
    reliability is unclear — see docs/data-hygiene.md. The `owing_money`
    tag is gone for the same reason.

    Note: under the Active++ filter, churned rows are not imported at
    all, so the `churned` tag effectively cannot fire. Kept in the
    logic defensively in case the filter ever widens.
    """
    nps_l = (nps_standing or "").strip().lower()
    tags: list[str] = []

    if nps_l == "detractor / at risk":
        tags.append("at_risk")
        tags.append("detractor")
    if nps_l == "promoter":
        tags.append("promoter")
    if status == "churned":
        tags.append("churned")
    if is_aus:
        tags.append("aus")

    return tags


@dataclass(frozen=True)
class OwnerParse:
    """Result of parsing a raw Owner cell."""

    team_email: str | None
    is_clean_match: bool
    raw: str | None


def parse_owner(raw: Any) -> OwnerParse:
    """Parse a messy Owner cell into (mapped_email, clean?, raw).

    Strategy: find the earliest first-name substring from OWNER_FIRST_NAMES.
    If that substring is the whole trimmed string (after lowering), it's
    a clean match — no raw_owner metadata needed. Otherwise we still
    assign to that team member but retain the raw string for audit.

    Strings with no matching first-name substring (N/A, Aleks, blanks)
    return team_email=None and the caller produces no assignment.
    """
    if raw is None:
        return OwnerParse(None, True, None)
    text = str(raw).strip()
    if not text:
        return OwnerParse(None, True, None)
    lower = text.lower()
    earliest_pos = len(lower) + 1
    earliest_email: str | None = None
    earliest_name: str | None = None
    for name, email in OWNER_FIRST_NAMES.items():
        pos = lower.find(name)
        if 0 <= pos < earliest_pos:
            earliest_pos = pos
            earliest_email = email
            earliest_name = name
    if earliest_email is None:
        return OwnerParse(None, False, text)
    is_clean = lower == earliest_name
    return OwnerParse(earliest_email, is_clean, text)


def _cell_to_date(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def build_client_payload(
    sheet_row: dict[str, Any],
    *,
    country: str,
    seeded_at_iso: str,
) -> dict[str, Any] | None:
    """Build the `clients` row payload, or return None if email is missing."""
    email = normalize_email(sheet_row.get("client emails"))
    if not email:
        return None

    full_name = (sheet_row.get("customer name") or "").strip() or None
    if not full_name:
        return None

    phone_raw = sheet_row.get("client phone no.")
    phone = str(phone_raw).strip() if phone_raw not in (None, "") else None

    slack_user_id_raw = sheet_row.get("slack user id")
    slack_user_id = (
        str(slack_user_id_raw).strip() if slack_user_id_raw not in (None, "") else None
    )

    start_date = _cell_to_date(sheet_row.get("date"))
    status = derive_status(sheet_row.get("status"))

    nps_raw = sheet_row.get("nps standing")
    nps_standing = str(nps_raw).strip() if nps_raw not in (None, "") else None

    tags = derive_tags(status, nps_standing, is_aus=(country == "AUS"))

    owner_raw = sheet_row.get("owner")
    owner_raw_str = str(owner_raw).strip() if owner_raw not in (None, "") else None

    # Revenue fields and Standing are both excluded. Revenue because the
    # sheet values are stale; Standing because its reliability is unclear
    # and its downstream tag effects (owing_money, at_risk) would have
    # carried that uncertainty into agent behavior.
    # See docs/data-hygiene.md.
    metadata = {
        "seed_source": "financial_master_jan26",
        "seeded_at": seeded_at_iso,
        "country": country,
        "nps_standing": nps_standing,
        "owner_raw": owner_raw_str,
    }

    return {
        "email": email,
        "full_name": full_name,
        "phone": phone,
        "slack_user_id": slack_user_id,
        "start_date": start_date.isoformat() if start_date else None,
        "status": status,
        "tags": tags,
        "metadata": metadata,
    }


def build_channel_payload(
    sheet_row: dict[str, Any], client_email: str, client_full_name: str
) -> dict[str, Any] | None:
    raw = sheet_row.get("slack channel id")
    if raw is None or not str(raw).strip():
        return None
    return {
        "slack_channel_id": str(raw).strip(),
        "name": client_full_name,
        "client_email": client_email,
        "is_private": True,
        "is_archived": False,
        "ella_enabled": False,
        "metadata": {},
    }


def build_assignment_payload(
    sheet_row: dict[str, Any], client_email: str
) -> dict[str, Any] | None:
    """Produce an assignment payload or None.

    Returns a dict keyed by `team_email` rather than team_member_id so
    the caller can resolve ids once at the top. Metadata carries
    raw_owner when the match wasn't clean.
    """
    parse = parse_owner(sheet_row.get("owner"))
    if parse.team_email is None:
        return None
    metadata: dict[str, Any] = {}
    if not parse.is_clean_match and parse.raw is not None:
        metadata["raw_owner"] = parse.raw
    return {
        "team_email": parse.team_email,
        "client_email": client_email,
        "role": "primary_csm",
        "metadata": metadata,
    }


# ---------------------------------------------------------------------------
# Sheet IO
# ---------------------------------------------------------------------------


def _headers_to_index(header_row: tuple) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key:
            idx[key] = i
    return idx


@dataclass
class SheetRow:
    """One row from a tab — raw sheet values + provenance."""

    values: dict[str, Any]
    country: str
    tab: str
    row_number: int


def load_sheet_rows(xlsx_path: Path) -> tuple[list[SheetRow], int]:
    """Return (rows, aggregate_rows_filtered).

    Aggregate / summary rows at the bottom of each tab (`TOTALS`,
    `UF Collection Rate`, etc.) are filtered here so they don't pollute
    the skipped-missing-email report. The filter is a known-label set
    (see AGGREGATE_ROW_LABELS); real people with blank Customer Name
    cells don't exist and real people with email-missing are preserved.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rows: list[SheetRow] = []
    aggregate_filtered = 0
    for tab, country in ((USA_TAB, "USA"), (AUS_TAB, "AUS")):
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            continue
        headers = _headers_to_index(raw_rows[0])
        for row_number, raw in enumerate(raw_rows[1:], start=2):
            values = {key: raw[idx] for key, idx in headers.items() if idx < len(raw)}
            customer_name = values.get("customer name")
            if not customer_name:
                continue
            if is_aggregate_row(customer_name):
                aggregate_filtered += 1
                continue
            rows.append(
                SheetRow(values=values, country=country, tab=tab, row_number=row_number)
            )
    return rows, aggregate_filtered


def locate_xlsx(explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit)
        if not p.exists():
            sys.exit(f"ERROR: {p} does not exist.")
        return p
    DEFAULT_INPUT_DIR.mkdir(parents=True, exist_ok=True)
    candidates = sorted(DEFAULT_INPUT_DIR.glob("*.xlsx"))
    if not candidates:
        sys.exit(
            f"ERROR: no .xlsx files in {DEFAULT_INPUT_DIR}. "
            "Drop the Financial Master Sheet there or pass --input."
        )
    if len(candidates) > 1:
        sys.exit(
            f"ERROR: multiple .xlsx files in {DEFAULT_INPUT_DIR}: "
            f"{[c.name for c in candidates]}. Pass --input to pick one."
        )
    return candidates[0]


# ---------------------------------------------------------------------------
# Report building
# ---------------------------------------------------------------------------


@dataclass
class ArchivalPlan:
    """Clients in the DB that are not in the new proposed set.

    Applied with a cascade: clients.archived_at = now, matching
    slack_channels.is_archived = true, active client_team_assignments
    ended with unassigned_at = now. No deletes.
    """

    clients_to_archive: list[dict[str, Any]] = field(default_factory=list)
    expected_channel_archivals: int = 0
    expected_assignment_unassignments: int = 0


@dataclass
class DryRunReport:
    total_sheet_rows: int = 0
    aggregate_rows_filtered: int = 0
    proposed_clients: list[dict[str, Any]] = field(default_factory=list)
    proposed_channels: list[dict[str, Any]] = field(default_factory=list)
    proposed_assignments: list[dict[str, Any]] = field(default_factory=list)
    skipped_missing_email: list[tuple[str, str, int]] = field(default_factory=list)
    excluded_by_view: list[tuple[str, str, str, int, str]] = field(default_factory=list)
    unmapped_owner_values: dict[str, int] = field(default_factory=dict)
    messy_owner_mappings: dict[str, tuple[str, int]] = field(default_factory=dict)
    sheet_email_duplicates: dict[str, list[tuple[str, str, int]]] = field(
        default_factory=dict
    )
    db_email_collisions: list[str] = field(default_factory=list)
    archival_plan: ArchivalPlan = field(default_factory=ArchivalPlan)


def build_report(
    rows: list[SheetRow],
    existing_emails_in_db: set[str],
    *,
    aggregate_rows_filtered: int = 0,
) -> DryRunReport:
    now_iso = datetime.now(timezone.utc).date().isoformat()
    report = DryRunReport(
        total_sheet_rows=len(rows) + aggregate_rows_filtered,
        aggregate_rows_filtered=aggregate_rows_filtered,
    )
    seen_emails: dict[str, list[tuple[str, str, int]]] = {}

    for row in rows:
        payload = build_client_payload(
            row.values, country=row.country, seeded_at_iso=now_iso
        )
        if payload is None:
            report.skipped_missing_email.append(
                (str(row.values.get("customer name") or "").strip(), row.tab, row.row_number)
            )
            continue

        # Active++ working-view filter — canonical V1 "what counts as a
        # client" rule. Excluded rows surface in the report but do not
        # become proposed inserts, channels, or assignments.
        if not is_in_active_plus_plus_view(payload["status"], row.country):
            report.excluded_by_view.append((
                payload["email"],
                payload["full_name"],
                row.tab,
                row.row_number,
                payload["status"],
            ))
            continue

        email = payload["email"]
        seen_emails.setdefault(email, []).append((payload["full_name"], row.tab, row.row_number))
        report.proposed_clients.append(payload)

        chan = build_channel_payload(row.values, email, payload["full_name"])
        if chan is not None:
            report.proposed_channels.append(chan)

        parse = parse_owner(row.values.get("owner"))
        if parse.team_email is not None:
            if not parse.is_clean_match and parse.raw is not None:
                report.messy_owner_mappings[parse.raw] = (
                    parse.team_email,
                    report.messy_owner_mappings.get(parse.raw, (parse.team_email, 0))[1] + 1,
                )
            assignment = build_assignment_payload(row.values, email)
            if assignment is not None:
                report.proposed_assignments.append(assignment)
        else:
            owner_raw = row.values.get("owner")
            if owner_raw is not None and str(owner_raw).strip():
                text = str(owner_raw).strip()
                report.unmapped_owner_values[text] = (
                    report.unmapped_owner_values.get(text, 0) + 1
                )

    for email, occurrences in seen_emails.items():
        if len(occurrences) > 1:
            report.sheet_email_duplicates[email] = occurrences

    report.db_email_collisions = sorted(
        {p["email"] for p in report.proposed_clients} & existing_emails_in_db
    )
    return report


def render_report(report: DryRunReport, *, sample_size: int = 5) -> str:
    """Render the dry-run report to a string."""
    lines: list[str] = []
    lines.append("=" * 72)
    lines.append("CLIENTS IMPORTER — DRY RUN REPORT")
    lines.append("=" * 72)
    lines.append("")
    lines.append(f"Total sheet rows with Customer Name:   {report.total_sheet_rows}")
    lines.append(f"Aggregate rows filtered (sheet totals):{report.aggregate_rows_filtered}")
    lines.append(f"Rows skipped (missing email):          {len(report.skipped_missing_email)}")
    lines.append(f"Rows excluded (Active++ filter):       {len(report.excluded_by_view)}")
    lines.append(f"Proposed clients inserts/updates:      {len(report.proposed_clients)}")
    lines.append(f"Proposed slack_channels inserts:       {len(report.proposed_channels)}")
    lines.append(f"Proposed client_team_assignments:      {len(report.proposed_assignments)}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("SKIPPED — missing email")
    lines.append("-" * 72)
    if not report.skipped_missing_email:
        lines.append("(none)")
    for name, tab, row_number in report.skipped_missing_email:
        lines.append(f"  {tab}  row {row_number:>3}  {name}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("EXCLUDED BY ACTIVE++ FILTER — not imported")
    lines.append("-" * 72)
    if not report.excluded_by_view:
        lines.append("(none)")
    else:
        status_totals: Counter[str] = Counter(st for _, _, _, _, st in report.excluded_by_view)
        lines.append("  by status:")
        for status, count in status_totals.most_common():
            lines.append(f"    {count:>3}  {status}")
        lines.append("")
        lines.append(f"  (showing first 15 of {len(report.excluded_by_view)})")
        for email, name, tab, row_number, status in report.excluded_by_view[:15]:
            lines.append(f"    {tab}  row {row_number:>3}  {status:<8}  {name}  <{email}>")
    lines.append("")

    lines.append("-" * 72)
    lines.append("UNMAPPED OWNER VALUES — no assignment produced")
    lines.append("-" * 72)
    if not report.unmapped_owner_values:
        lines.append("(none)")
    for value, count in sorted(report.unmapped_owner_values.items(), key=lambda kv: -kv[1]):
        lines.append(f"  {count:>3} rows  owner={value!r}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("MESSY OWNER MAPPINGS — raw_owner saved to metadata")
    lines.append("-" * 72)
    if not report.messy_owner_mappings:
        lines.append("(none)")
    for raw, (email, count) in sorted(
        report.messy_owner_mappings.items(), key=lambda kv: -kv[1][1]
    ):
        lines.append(f"  {count:>3} rows  {raw!r:<30}  ->  {email}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("EMAIL DUPLICATES WITHIN THE SHEET")
    lines.append("-" * 72)
    if not report.sheet_email_duplicates:
        lines.append("(none)")
    for email, occurrences in report.sheet_email_duplicates.items():
        lines.append(f"  {email}")
        for name, tab, row_number in occurrences:
            lines.append(f"      {tab}  row {row_number:>3}  {name}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("EMAIL COLLISIONS WITH EXISTING DB ROWS")
    lines.append("-" * 72)
    if not report.db_email_collisions:
        lines.append("(none)")
    for email in report.db_email_collisions:
        lines.append(f"  {email}")
    lines.append("")

    lines.append("-" * 72)
    lines.append("PROPOSED ARCHIVALS — DB rows not in the new Active++ set")
    lines.append("-" * 72)
    plan = report.archival_plan
    if not plan.clients_to_archive:
        lines.append("(none)")
    else:
        lines.append(f"  {len(plan.clients_to_archive)} clients will be soft-archived (archived_at = now)")
        lines.append(f"  {plan.expected_channel_archivals} slack_channels will be marked is_archived = true")
        lines.append(f"  {plan.expected_assignment_unassignments} client_team_assignments will be unassigned (unassigned_at = now)")
        lines.append("")
        arch_status_totals = Counter(row.get("status") or "(null)" for row in plan.clients_to_archive)
        lines.append("  clients to archive, by status:")
        for status, count in arch_status_totals.most_common():
            lines.append(f"    {count:>3}  {status}")
        lines.append("")
        shown = plan.clients_to_archive[:10]
        if shown:
            lines.append(f"  sample (first {len(shown)} of {len(plan.clients_to_archive)}):")
            for row in shown:
                lines.append(f"    {row['status']:<8}  {row['full_name']}  <{row['email']}>")
    lines.append("")

    lines.append("-" * 72)
    lines.append(f"SAMPLE — {sample_size} random clients + guaranteed AUS sample")
    lines.append("-" * 72)

    if report.proposed_clients:
        rng = random.Random(42)
        sample = rng.sample(
            report.proposed_clients,
            k=min(sample_size, len(report.proposed_clients)),
        )
        aus_in_sample = any(p["metadata"]["country"] == "AUS" for p in sample)
        if not aus_in_sample:
            aus_rows = [p for p in report.proposed_clients if p["metadata"]["country"] == "AUS"]
            if aus_rows:
                sample.append(rng.choice(aus_rows))
        for p in sample:
            lines.append("")
            lines.append(f"  email:        {p['email']}")
            lines.append(f"  full_name:    {p['full_name']}")
            lines.append(f"  phone:        {p['phone']}")
            lines.append(f"  slack_user_id:{p['slack_user_id']}")
            lines.append(f"  start_date:   {p['start_date']}")
            lines.append(f"  status:       {p['status']}")
            lines.append(f"  tags:         {p['tags']}")
            lines.append(f"  metadata:     {json.dumps(p['metadata'], default=str)}")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------


def resolve_team_member_ids(client) -> dict[str, str]:
    resp = client.table("team_members").select("id,email").execute()
    return {row["email"]: row["id"] for row in resp.data}


def fetch_existing_client_emails(client, emails: Iterable[str]) -> dict[str, dict[str, Any]]:
    """Return {email: {id, metadata}} for rows with archived_at is null."""
    email_list = list({e for e in emails if e})
    if not email_list:
        return {}
    resp = (
        client.table("clients")
        .select("id,email,metadata")
        .is_("archived_at", "null")
        .in_("email", email_list)
        .execute()
    )
    return {row["email"]: row for row in resp.data}


def apply_clients(
    client, proposed: list[dict[str, Any]]
) -> tuple[dict[str, str], int, int]:
    """Insert or update clients. Returns (email->id, inserts, updates)."""
    existing = fetch_existing_client_emails(client, (p["email"] for p in proposed))
    email_to_id: dict[str, str] = {}
    inserts = 0
    updates = 0

    for p in proposed:
        row_insert = {
            "email": p["email"],
            "full_name": p["full_name"],
            "phone": p["phone"],
            "slack_user_id": p["slack_user_id"],
            "start_date": p["start_date"],
            "status": p["status"],
            "tags": p["tags"],
        }
        if p["email"] in existing:
            existing_row = existing[p["email"]]
            merged_metadata = (existing_row.get("metadata") or {}) | p["metadata"]
            update_payload = dict(row_insert)
            update_payload["metadata"] = merged_metadata
            resp = (
                client.table("clients")
                .update(update_payload)
                .eq("id", existing_row["id"])
                .execute()
            )
            email_to_id[p["email"]] = existing_row["id"]
            updates += 1
        else:
            row_insert["metadata"] = p["metadata"]
            resp = client.table("clients").insert(row_insert).execute()
            email_to_id[p["email"]] = resp.data[0]["id"]
            inserts += 1

    return email_to_id, inserts, updates


def apply_channels(
    client, proposed: list[dict[str, Any]], email_to_client_id: dict[str, str]
) -> int:
    count = 0
    for p in proposed:
        client_id = email_to_client_id.get(p["client_email"])
        if client_id is None:
            continue
        payload = {
            "slack_channel_id": p["slack_channel_id"],
            "name": p["name"],
            "client_id": client_id,
            "is_private": p["is_private"],
            "is_archived": p["is_archived"],
            "ella_enabled": p["ella_enabled"],
        }
        client.table("slack_channels").upsert(
            payload, on_conflict="slack_channel_id"
        ).execute()
        count += 1
    return count


def apply_assignments(
    client,
    proposed: list[dict[str, Any]],
    email_to_client_id: dict[str, str],
    team_email_to_id: dict[str, str],
) -> int:
    count = 0
    for p in proposed:
        client_id = email_to_client_id.get(p["client_email"])
        team_member_id = team_email_to_id.get(p["team_email"])
        if client_id is None or team_member_id is None:
            continue
        payload = {
            "client_id": client_id,
            "team_member_id": team_member_id,
            "role": p["role"],
            "metadata": p["metadata"],
        }
        client.table("client_team_assignments").upsert(
            payload,
            on_conflict="client_id,team_member_id,role",
            ignore_duplicates=True,
        ).execute()
        count += 1
    return count


def compute_archival_plan(client, proposed_emails: set[str]) -> ArchivalPlan:
    """Find currently-active clients not in the new proposed set.

    These are the casualties of the filter revision: rows that were
    valid under an earlier rule but aren't in the sheet's current
    Active++ view. On --apply they get soft-archived with the cascade.
    """
    resp = (
        client.table("clients")
        .select("id,email,full_name,status")
        .is_("archived_at", "null")
        .execute()
    )
    existing = resp.data or []
    to_archive = [row for row in existing if row["email"] not in proposed_emails]

    plan = ArchivalPlan(clients_to_archive=to_archive)
    if not to_archive:
        return plan

    ids = [row["id"] for row in to_archive]
    ch_resp = (
        client.table("slack_channels")
        .select("id")
        .in_("client_id", ids)
        .eq("is_archived", False)
        .execute()
    )
    plan.expected_channel_archivals = len(ch_resp.data or [])

    a_resp = (
        client.table("client_team_assignments")
        .select("id")
        .in_("client_id", ids)
        .is_("unassigned_at", "null")
        .execute()
    )
    plan.expected_assignment_unassignments = len(a_resp.data or [])

    return plan


def apply_archival(client, plan: ArchivalPlan) -> tuple[int, int, int]:
    """Execute the archival plan. Returns (clients, channels, assignments)."""
    if not plan.clients_to_archive:
        return 0, 0, 0

    now_iso = datetime.now(timezone.utc).isoformat()
    ids = [row["id"] for row in plan.clients_to_archive]

    clients_resp = (
        client.table("clients")
        .update({"archived_at": now_iso})
        .in_("id", ids)
        .execute()
    )
    clients_count = len(clients_resp.data or [])

    ch_resp = (
        client.table("slack_channels")
        .update({"is_archived": True})
        .in_("client_id", ids)
        .eq("is_archived", False)
        .execute()
    )
    channels_count = len(ch_resp.data or [])

    a_resp = (
        client.table("client_team_assignments")
        .update({"unassigned_at": now_iso})
        .in_("client_id", ids)
        .is_("unassigned_at", "null")
        .execute()
    )
    assignments_count = len(a_resp.data or [])

    return clients_count, channels_count, assignments_count


def apply_log_breakdowns(client) -> str:
    """Render status / journey_stage / tag breakdowns over all active clients.

    Called after --apply so the log captures what actually landed. Fetches
    status, journey_stage, and tags columns in a single SELECT and tallies
    in Python — faster than three GROUP BY roundtrips via PostgREST for
    the V1-scale row count.
    """
    resp = (
        client.table("clients")
        .select("status,journey_stage,tags")
        .is_("archived_at", "null")
        .execute()
    )
    rows = resp.data or []
    total = len(rows)

    status_counts: Counter[str] = Counter(r.get("status") or "(null)" for r in rows)
    journey_counts: Counter[str] = Counter(
        r.get("journey_stage") or "(null)" for r in rows
    )
    tag_counts: Counter[str] = Counter()
    for r in rows:
        for t in r.get("tags") or []:
            tag_counts[t] += 1

    def _format_counter(title: str, counter: Counter[str]) -> list[str]:
        lines = [title, "-" * len(title)]
        if not counter:
            lines.append("(none)")
            return lines
        width = max(len(str(k)) for k in counter)
        for key, count in counter.most_common():
            lines.append(f"  {str(key).ljust(width)}  {count}")
        return lines

    sections: list[str] = []
    sections.append(f"Breakdowns across {total} active clients:")
    sections.append("")
    sections.extend(_format_counter("status", status_counts))
    sections.append("")
    sections.extend(_format_counter("journey_stage", journey_counts))
    sections.append("")
    sections.extend(_format_counter("tags (clients counted once per tag)", tag_counts))
    return "\n".join(sections)


def write_log(report_text: str, apply_text: str | None = None) -> Path:
    IMPORT_LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = IMPORT_LOG_DIR / f"import_{ts}.log"
    body = report_text
    if apply_text:
        body = body + "\n" + apply_text
    path.write_text(body)
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", help="Path to the XLSX. Auto-detects in data/client_seed/ otherwise.")
    parser.add_argument("--apply", action="store_true", help="Write changes. Without this flag, dry-run only.")
    args = parser.parse_args(argv)

    xlsx = locate_xlsx(args.input)
    rows, aggregate_filtered = load_sheet_rows(xlsx)

    db = get_client()
    existing_emails_db = set(
        fetch_existing_client_emails(db, (
            normalize_email(r.values.get("client emails"))
            for r in rows if r.values.get("client emails")
        )).keys()
    )

    report = build_report(rows, existing_emails_db, aggregate_rows_filtered=aggregate_filtered)
    report.archival_plan = compute_archival_plan(
        db, proposed_emails={p["email"] for p in report.proposed_clients}
    )
    report_text = render_report(report)
    print(report_text)

    if not args.apply:
        print("Dry run only — no changes written. Re-run with --apply to commit.")
        return 0

    print("-" * 72)
    print("APPLYING...")
    print("-" * 72)
    team_email_to_id = resolve_team_member_ids(db)

    email_to_client_id, c_inserts, c_updates = apply_clients(db, report.proposed_clients)
    ch_count = apply_channels(db, report.proposed_channels, email_to_client_id)
    as_count = apply_assignments(
        db, report.proposed_assignments, email_to_client_id, team_email_to_id
    )

    arch_clients, arch_channels, arch_assignments = apply_archival(db, report.archival_plan)

    apply_text = (
        "APPLY SUMMARY\n"
        f"  clients inserts:                    {c_inserts}\n"
        f"  clients updates:                    {c_updates}\n"
        f"  slack_channels upserts:             {ch_count}\n"
        f"  client_team_assignments upserts:    {as_count}\n"
        f"  clients archived (soft):            {arch_clients}\n"
        f"  slack_channels archived (cascade):  {arch_channels}\n"
        f"  assignments ended (cascade):        {arch_assignments}\n"
    )
    print("\n" + apply_text)

    discrepancies = _render_discrepancies(report, arch_clients, arch_channels, arch_assignments)
    print(discrepancies)

    breakdowns = apply_log_breakdowns(db)
    print("\n" + breakdowns)

    log_body = apply_text + "\n" + discrepancies + "\n\n" + breakdowns
    log_path = write_log(report_text, log_body)
    print(f"\nLog: {log_path}")
    return 0


def _render_discrepancies(
    report: DryRunReport,
    actual_archived_clients: int,
    actual_archived_channels: int,
    actual_archived_assignments: int,
) -> str:
    """Compare dry-run-predicted counts to what --apply actually wrote."""
    lines: list[str] = []
    lines.append("DISCREPANCY CHECKS")
    lines.append("-" * 72)

    expected_archived = len(report.archival_plan.clients_to_archive)
    expected_channel_archivals = report.archival_plan.expected_channel_archivals
    expected_assignment_unassignments = report.archival_plan.expected_assignment_unassignments

    checks = [
        ("clients archived",               expected_archived, actual_archived_clients),
        ("slack_channels archived",        expected_channel_archivals, actual_archived_channels),
        ("assignments ended",              expected_assignment_unassignments, actual_archived_assignments),
    ]
    any_flag = False
    for label, expected, actual in checks:
        mark = "OK " if expected == actual else "!! "
        if expected != actual:
            any_flag = True
        lines.append(f"  {mark}  {label:<30}  expected {expected:>4}  actual {actual:>4}")
    if not any_flag:
        lines.append("  (all dry-run predictions matched the applied counts)")
    return "\n".join(lines)


if __name__ == "__main__":
    sys.exit(main())
